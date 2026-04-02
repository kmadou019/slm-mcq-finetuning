"""
Génère un tableau Excel comparant la couverture des sections
du papier "Model Cards for Model Reporting" (Mitchell et al., 2019)
pour les modèles utilisés dans le projet slm-mcq-finetuning.

Légende :
  ✅  Section clairement présente
  ⚠️  Section partiellement présente
  ❌  Section absente
  —   Sans objet (modèle custom / fine-tune local, pas de HF card)
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ---------------------------------------------------------------------------
# Données
# ---------------------------------------------------------------------------

SECTIONS = [
    "Model Details",
    "Intended Use",
    "Factors\n(groupes démo.)",
    "Métriques",
    "Données\névaluation",
    "Données\nentraînement",
    "Analyses quant.\ndésagrégées",
    "Considérations\néthiques",
    "Mises en garde\n& Recommand.",
]

# (nom_affichage, hf_id, groupe, [score par section])
# Score : ✅ / ⚠️ / ❌ / —
MODELS = [
    # ── Référence (exemples du papier / discussion) ──────────────────────
    ("BLOOM",            "bigscience/bloom",                  "Référence (papier)",
     ["✅","✅","⚠️","✅","✅","✅","⚠️","✅","✅"]),
    ("Llama-2-70B Chat", "meta-llama/Llama-2-70b-chat-hf",   "Référence (papier)",
     ["✅","✅","⚠️","✅","✅","⚠️","⚠️","✅","✅"]),
    ("Gemma-2-9B",       "google/gemma-2-9b",                 "Référence (papier)",
     ["✅","✅","❌","✅","⚠️","⚠️","❌","⚠️","⚠️"]),

    # ── Projet – CSVs évalués ─────────────────────────────────────────────
    ("Llama-3.1-8B",     "meta-llama/Llama-3.1-8B",          "Projet (CSVs)",
     ["✅","✅","❌","✅","⚠️","⚠️","❌","✅","⚠️"]),
    ("OpenBioLLM-8B",    "aaditya/OpenBioLLM-Llama3-8B",     "Projet (CSVs)",
     ["✅","✅","❌","✅","✅","⚠️","❌","⚠️","⚠️"]),
    ("MedGemma-4B",      "google/medgemma-4b",                "Projet (CSVs)",
     ["✅","✅","⚠️","✅","✅","⚠️","⚠️","✅","✅"]),
    ("MedGemma-27B",     "google/medgemma-27b",               "Projet (CSVs)",
     ["✅","✅","⚠️","✅","✅","⚠️","⚠️","✅","✅"]),
    ("Qwen3-0.6B",       "Qwen/Qwen3-0.6B",                  "Projet (CSVs)",
     ["✅","✅","❌","✅","⚠️","⚠️","❌","⚠️","⚠️"]),
    ("Qwen3-1.7B",       "Qwen/Qwen3-1.7B",                  "Projet (CSVs)",
     ["✅","✅","❌","✅","⚠️","⚠️","❌","⚠️","⚠️"]),
    ("Qwen3-4B",         "Qwen/Qwen3-4B",                    "Projet (CSVs)",
     ["✅","✅","❌","✅","⚠️","⚠️","❌","⚠️","⚠️"]),
    ("Qwen3-8B",         "Qwen/Qwen3-8B",                    "Projet (CSVs)",
     ["✅","✅","❌","✅","⚠️","⚠️","❌","⚠️","⚠️"]),
    ("Mistral-7B",       "mistralai/Mistral-7B-v0.1",        "Projet (CSVs)",
     ["✅","⚠️","❌","✅","⚠️","❌","❌","❌","❌"]),
    ("EuroLLM-9B",       "EuroLLM/EuroLLM-9B-Instruct",      "Projet (CSVs)",
     ["✅","✅","⚠️","✅","✅","⚠️","❌","⚠️","⚠️"]),
    ("Apertus-8B",       "BramVanroy/apertus-8B",             "Projet (CSVs)",
     ["✅","✅","❌","⚠️","⚠️","⚠️","❌","⚠️","⚠️"]),

    # ── Projet – fine-tunes locaux (pas de HF card publique) ─────────────
    ("Qwen3-4B pdapt-slerp", "fine-tune local",              "Projet (fine-tunes)",
     ["—","—","—","—","—","—","—","—","—"]),
    ("Qwen3-8B pdapt-slerp", "fine-tune local",              "Projet (fine-tunes)",
     ["—","—","—","—","—","—","—","—","—"]),

    # ── Benchmarks (thinking vs no_thinking) ─────────────────────────────
    ("Qwen3-14B",            "Qwen/Qwen3-14B",               "Benchmarks",
     ["✅","✅","❌","✅","⚠️","⚠️","❌","⚠️","⚠️"]),
    ("Magistral-24B",        "mistralai/Magistral-Small-24B","Benchmarks",
     ["✅","✅","❌","✅","⚠️","❌","❌","⚠️","⚠️"]),
    ("Mistral-Small-3.1-24B","mistralai/Mistral-Small-3.1-24B","Benchmarks",
     ["✅","✅","❌","✅","⚠️","⚠️","❌","⚠️","⚠️"]),
    ("Nemotron-30B",         "nvidia/Nemotron-3-Nano-30B",   "Benchmarks",
     ["✅","✅","❌","✅","✅","⚠️","❌","✅","⚠️"]),
    ("DeepSeek-R1-14B",      "deepseek-ai/DeepSeek-R1-Distill-Qwen-14B","Benchmarks",
     ["✅","✅","❌","✅","✅","⚠️","❌","⚠️","⚠️"]),
    ("Qwen2.5-14B",          "Qwen/Qwen2.5-14B",             "Benchmarks",
     ["✅","✅","❌","✅","⚠️","⚠️","❌","⚠️","⚠️"]),
    ("QwQ-32B",              "Qwen/QwQ-32B",                 "Benchmarks",
     ["✅","✅","❌","✅","⚠️","❌","❌","❌","⚠️"]),
    ("Qwen2.5-32B",          "Qwen/Qwen2.5-32B",             "Benchmarks",
     ["✅","✅","❌","✅","⚠️","⚠️","❌","⚠️","⚠️"]),
    ("Phi-4-14B",            "microsoft/phi-4",              "Benchmarks",
     ["✅","✅","❌","✅","✅","⚠️","❌","✅","✅"]),
    ("Phi-4-Reasoning-14B",  "microsoft/phi-4-reasoning",    "Benchmarks",
     ["✅","✅","❌","✅","⚠️","⚠️","❌","✅","⚠️"]),
]

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

FILL = {
    "✅":  make_fill("92D050"),   # vert
    "⚠️": make_fill("FFD966"),   # jaune
    "❌":  make_fill("FF6B6B"),   # rouge
    "—":   make_fill("D9D9D9"),   # gris
}
SYMBOL_LABEL = {
    "✅":  "✅  Présent",
    "⚠️": "⚠️  Partiel",
    "❌":  "❌  Absent",
    "—":   "—   N/A",
}

HEADER_FILL   = make_fill("1F4E79")
GROUP_FILLS   = {
    "Référence (papier)":   make_fill("BDD7EE"),
    "Projet (CSVs)":        make_fill("E2EFDA"),
    "Projet (fine-tunes)":  make_fill("FFF2CC"),
    "Benchmarks":           make_fill("FCE4D6"),
}
WHITE_FONT    = Font(color="FFFFFF", bold=True, size=10)
BOLD_FONT     = Font(bold=True, size=10)
NORMAL_FONT   = Font(size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

thick = Side(style="medium", color="555555")
THICK_BORDER = Border(left=thick, right=thick, top=thick, bottom=thick)

# ---------------------------------------------------------------------------
# Construction du workbook
# ---------------------------------------------------------------------------

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Model Card Coverage"

N_SECTIONS = len(SECTIONS)

# ── Ligne 1 : titre principal ────────────────────────────────────────────────
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + N_SECTIONS)
title_cell = ws.cell(row=1, column=1,
    value='Couverture des sections "Model Cards for Model Reporting" (Mitchell et al., 2019)')
title_cell.fill   = HEADER_FILL
title_cell.font   = Font(color="FFFFFF", bold=True, size=13)
title_cell.alignment = CENTER
ws.row_dimensions[1].height = 28

# ── Ligne 2 : en-têtes colonnes ──────────────────────────────────────────────
headers = ["Modèle", "HuggingFace ID", "Groupe"] + SECTIONS
for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=h)
    cell.fill      = HEADER_FILL
    cell.font      = WHITE_FONT
    cell.alignment = CENTER
    cell.border    = THIN_BORDER
ws.row_dimensions[2].height = 50

# ── Lignes données ────────────────────────────────────────────────────────────
prev_group = None
row_idx = 3
for (name, hf_id, group, scores) in MODELS:
    # Séparateur visuel entre groupes
    if group != prev_group:
        prev_group = group
        ws.row_dimensions[row_idx].height = 6
        sep_fill = make_fill("444444")
        for c in range(1, 4 + N_SECTIONS):
            ws.cell(row=row_idx, column=c).fill = sep_fill
        row_idx += 1

    gfill = GROUP_FILLS.get(group, make_fill("FFFFFF"))

    # Colonne modèle
    c = ws.cell(row=row_idx, column=1, value=name)
    c.fill = gfill; c.font = BOLD_FONT; c.alignment = LEFT; c.border = THIN_BORDER

    # Colonne HF ID
    c = ws.cell(row=row_idx, column=2, value=hf_id)
    c.fill = gfill; c.font = Font(size=9, italic=True, color="444444")
    c.alignment = LEFT; c.border = THIN_BORDER

    # Colonne groupe
    c = ws.cell(row=row_idx, column=3, value=group)
    c.fill = gfill; c.font = NORMAL_FONT; c.alignment = CENTER; c.border = THIN_BORDER

    # Colonnes sections
    for s_idx, score in enumerate(scores):
        c = ws.cell(row=row_idx, column=4 + s_idx, value=score)
        c.fill      = FILL.get(score, make_fill("FFFFFF"))
        c.font      = Font(size=11)
        c.alignment = CENTER
        c.border    = THIN_BORDER

    ws.row_dimensions[row_idx].height = 20
    row_idx += 1

# ── Légende ───────────────────────────────────────────────────────────────────
row_idx += 1
ws.cell(row=row_idx, column=1, value="Légende :").font = BOLD_FONT
row_idx += 1
for symbol, label in SYMBOL_LABEL.items():
    c = ws.cell(row=row_idx, column=1, value=label)
    c.fill = FILL[symbol]
    c.font = Font(size=10)
    c.alignment = LEFT
    c.border = THIN_BORDER
    ws.row_dimensions[row_idx].height = 18
    row_idx += 1

# ── Note méthodologique ───────────────────────────────────────────────────────
row_idx += 1
note = (
    "Note : évaluation basée sur l'analyse des HF model cards publiques (mars 2026). "
    "La colonne 'Analyses quant. désagrégées' correspond à l'évaluation par sous-groupes "
    "démographiques/intersectionnels, cœur du papier Mitchell et al. — quasi-absente sur HF."
)
ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3 + N_SECTIONS)
c = ws.cell(row=row_idx, column=1, value=note)
c.font = Font(size=9, italic=True, color="555555")
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[row_idx].height = 36

# ── Largeurs colonnes ─────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 26   # Modèle
ws.column_dimensions["B"].width = 38   # HF ID
ws.column_dimensions["C"].width = 22   # Groupe
for i in range(N_SECTIONS):
    col_letter = get_column_letter(4 + i)
    ws.column_dimensions[col_letter].width = 13

# ── Figer la ligne d'en-têtes ─────────────────────────────────────────────────
ws.freeze_panes = "A3"

# ---------------------------------------------------------------------------
# Sauvegarde
# ---------------------------------------------------------------------------

output_path = os.path.join(os.path.dirname(__file__), "model_card_coverage.xlsx")
wb.save(output_path)
print(f"Fichier sauvegardé : {output_path}")
